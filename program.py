import csv
import math
import re
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import os
import pdfkit

from jinja2 import Environment, FileSystemLoader
from functools import reduce
from datetime import datetime
from typing import List, Dict, Tuple
from openpyxl import Workbook
from openpyxl.styles import Side, Border, Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter


# не стал указывать типы передаваемых переменных в методы, так как изначально написал типизированный код


class Salary:
    """
    Класс для представления зарплаты
    """
    __currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, values: List[str]):
        """
        Инициализирует объект Salary

        Args:
            values (List[str]): Нижняя граница оклада, верхняя граница оклада, валюта оклада
        """
        [self.__salary_from, self.__salary_to, self.__salary_currency] = values

    def __float__(self) -> float:
        """Преобразует зарплату к float значению в рублях"""
        return (float(self.__salary_from) + float(self.__salary_to)) / 2 * self.__currency_to_rub[
            self.__salary_currency.upper()]


class Vacancy:
    """Класс для представления вакансии"""

    def __init__(self, row: List[str], title: List[str]):
        """
        Инициализирует объект класса Vacancy


        :param row: Строка с вакансией из csv файла
        :param title: Названия столбцов csv файла
        """
        self.__name = None
        self.__salary = None
        self.__area_name = None
        self.__published_at = None
        self.__salary_from = None
        self.__salary_to = None
        self.__salary_currency = None

        fields_cases = {
            'name': lambda value: self.__set_value('name', HelpMethods.delete_rubbish(value)),
            'salary_from': lambda value: self.__set_value('salary_from', HelpMethods.delete_rubbish(value)),
            'salary_to': lambda value: self.__set_value('salary_to', HelpMethods.delete_rubbish(value)),
            'salary_currency': lambda value: self.__set_value('salary_currency', HelpMethods.delete_rubbish(value)),
            'area_name': lambda value: self.__set_value('area_name', HelpMethods.delete_rubbish(value)),
            'published_at': lambda value: self.__set_value(
                'published_at',
                Vacancy.__get_date(HelpMethods.delete_rubbish(value))),
        }

        for i, field in enumerate(row):
            if title[i] not in fields_cases:
                continue

            fields_cases[title[i]](field)

        self.__set_salary()

    def get_salary(self) -> float:
        """Возращает зарплату в рублях для данной вакансии"""
        return float(self.__salary)

    def get_date(self) -> str:
        """Возвращает дату размещения вакансии"""
        return self.__published_at

    def get_area(self) -> str:
        """Возвращает город, в котором размещена данная вакансия"""
        return self.__area_name

    def is_suitible(self, name: str) -> bool:
        """
        Содержит ли в названии name

        :param name: Название вакансии
        """
        return self.__name.count(name) > 0

    def __set_value(self, key, value):
        """
        Метод для инициализации приватных полей объекта
        :param key: Название поля
        :param value: Значение поля
        """
        self.__dict__['_Vacancy__' + key] = value

    def __set_salary(self):
        """Инициализирует зарплату при инициализации объекта"""
        self.__salary = Salary([self.__salary_from, self.__salary_to, self.__salary_currency])

    @staticmethod
    def __get_date(date: str) -> str:
        """
        Вычисляет из строки год
        :param date: Дата
        :return: float: Год
        """
        return str(datetime.fromisoformat(date[:-2] + ":" + date[-2:]).year)


class DataSet:
    """Класс, представляющий набор данных обо всех вакансиях"""

    def __init__(self, file_name: str):
        """
        Инициализирует объект Dataset
        :param file_name: Название файла
        """
        self.__file_name = file_name
        self.__vacancies_objects: List[Vacancy] = []
        self.__title = None
        self.__vacancies_years = {}
        self.__vacancies_areas = {}
        self.__len = 0

        with open(file_name, mode='r', encoding='utf-8-sig') as vacancies:
            file_reader = csv.reader(vacancies, delimiter=",")
            has_title = False

            for row in file_reader:
                if not has_title:
                    self.__title = row
                    has_title = True
                    continue

                if row.count('') != 0 or len(row) < len(self.__title) - 1: continue

                self.__validate_vacancy(row)
                self.__len += 1

    def __validate_vacancy(self, row: List[str]):
        """
        Парсит валидную строку csv файла
        :param row: Строка
        """
        vacancy = Vacancy(row, self.__title)

        now_date = self.__vacancies_years.get(vacancy.get_date(), [])
        now_date.append(vacancy)
        self.__vacancies_years[vacancy.get_date()] = now_date

        now_area = self.__vacancies_areas.get(vacancy.get_area(), [])
        now_area.append(vacancy)
        self.__vacancies_areas[vacancy.get_area()] = now_area

    def get_vacancies_years(self, func=None) -> Dict[str, List[int]]:
        """
        Создает словарь с ключами-годами и значениями - массивами из зарплат в соответствии с фильтрующей
        функцией
        :param func: Фильтрующая функция
        :return: Словарь с массивами зарплат по годам
        """
        if func is None:
            return DataSet.get_structured_salaries(self.__vacancies_years)

        result = {}

        for year in self.__vacancies_years.keys():
            result[year] = []
            for vacancy in self.__vacancies_years[year]:
                if func(vacancy):
                    result[year].append(vacancy)

        return DataSet.get_structured_salaries(result)

    def get_vacancies_cities(self) -> Tuple[List[List[float]], List[List[int]]]:
        """
        Создает кортеж из листов с долями вакансий и уровнем зарплат по городам
        :return: Кортеж из листов с долями вакансий и уровнем зарплат по городам
        """
        cities_s = []
        fract = []

        for key, value in self.__vacancies_areas.items():

            percent = round(len(value) / self.__len, 4)
            if percent < 0.01:
                continue

            summ = 0
            for vacancy in self.__vacancies_areas[key]:
                summ += vacancy.get_salary()

            cities_s.append([key, math.floor(summ / len(value))])
            fract.append([key, percent])

        fract.sort(key=lambda x: x[1], reverse=True)
        cities_s.sort(key=lambda x: x[1], reverse=True)

        return fract, cities_s

    @staticmethod
    def get_structured_salaries(vacancies: Dict[str, list]) -> Dict[str, List[int]]:
        """
        Создает словарь с ключами-годами и значениями - массивами из зарплат
        :param vacancies: Датасет вакансий
        :return: Словарь с ключами-годами и значениями - массивами из зарплат
        """
        salaries = {}

        for i, year in enumerate(vacancies.keys()):

            summ = 0
            for vacancy in vacancies[year]:
                summ += vacancy.get_salary()

            salaries[year] = [
                math.floor(summ / len(vacancies[year])) if len(vacancies[year]) > 0 else 0,
                len(vacancies[year])
            ]

        return salaries


class InputConnect:
    """
    Класс, представляющий взаимодействие пользователся и программы

    Attributes:
        file_name (str): Название файла
        vacancy (str): Название профессии для фильтрации
        method (str): Метод вывода данных
    """

    def __init__(self):
        """
        Инициализирует объект класса InputConnect
        """
        self.file_name = None
        self.vacancy = None
        self.method = None

    def read_console(self):
        """
        Метод для чтения данных с консоли и их сохранения
        """
        # if input() == '':
        #     self.file_name = '../vacancies.csv'
        #     self.vacancy = 'Аналитик'
        #     self.method = 'Вакансии'
        # else:
        #     self.file_name = input("Введите название файла: ")
        #     self.vacancy = input("Введите название профессии: ")
        #     self.method = input("Вакансии или статистика: ")

        self.file_name = input("Введите название файла: ")
        self.vacancy = input("Введите название профессии: ")
        self.method = input("Вакансии или статистика: ")

    @staticmethod
    def write_console(s_all, s_filtered, fract, cities_s):
        """
        Метод для вывода вакансий в консоль
        :param s_all: Словарь с ключами-годами и значениями - массивами из зарплат
        :param s_filtered: Словарь с ключами-годами и значениями - массивами из зарплат для выбранной профессии
        :param fract: Доли вакансий по городам
        :param cities_s: Средние зарплаты по городам
        """
        InputConnect.__write_salaries(s_all)
        InputConnect.__write_salaries(s_filtered, ' для выбранной профессии')

        InputConnect.__write_salaries_cities(fract, cities_s)

    @staticmethod
    def __write_salaries(salaries: Dict[str, List[int]], sufix=''):
        """
        Выводит в консоль словарь с ключами-годами и значениями - массивами из зарплат
        :param salaries: Словарь с ключами-годами и значениями - массивами из зарплат
        :param sufix: Доп параметр для печати
        """
        s = f'Динамика уровня зарплат по годам{sufix}' + ': {'

        print(s, end='')

        for i, year in enumerate(salaries.keys()):
            if i != 0:
                print(', ', end='')

            print(f'{year}: {salaries[year][0]}', end='')

        print('}')

        s = f'Динамика количества вакансий по годам{sufix}' + ': {'

        print(s, end='')

        for i, year in enumerate(salaries.keys()):
            if i != 0:
                print(', ', end='')

            print(f'{year}: {salaries[year][1]}', end='')

        print('}')

    @staticmethod
    def __write_salaries_cities(fract: List[List[float]], cities_s: List[List[int]]):
        """
        Выводит уровень зарплат и доли вакансий по городам в консоль
        :param fract: Доля вакансий по городам
        :param cities_s: Средние зарплаты по городам
        """
        print('Уровень зарплат по городам (в порядке убывания): {', end='')
        for i, e in enumerate(cities_s[:10]):
            if i != 0:
                print(', ', end='')
            print(f"'{e[0]}': {e[1]}", end='')

        print('}')

        print('Доля вакансий по городам (в порядке убывания): {', end='')
        for i, e in enumerate(fract[:10]):
            if i != 0:
                print(', ', end='')
            print(f"'{e[0]}': {e[1]}", end='')

        print('}')


class HelpMethods:
    """
    Класс, содержащий в себе вспомогатильные функции которые могут быть переиспользованы
    """
    @staticmethod
    def quit_program(message):
        """
        Завершить программу, предварительно напечатав сообщение в консоль
        :param message: Сообщение
        """
        print(message)
        quit()

    @staticmethod
    def delete_rubbish(s: str) -> str:
        """
        Удаляет html теги и лишние пробелы из строки
        :param s: Строка для чистки
        :return: Очищенная строка
        """
        rubbish_html = re.compile('<.*?>')

        return ' '.join(re.sub(rubbish_html, '', s).split()).strip()


class report:
    """
    Класс для представления разлияных видов отчетов
    """
    def __init__(self, vacancy: str,
                 s_all: Dict[str, List[int]],
                 s_filtered: Dict[str, List[int]],
                 fract: List[List[float]],
                 cities_s: List[List[int]]):
        """
        Инициализирует объект класса report
        :param vacancy: Вакансия, по которой была произведена фильтрация
        :param s_all: Словарь с ключами-годами и значениями - массивами из зарплат
        :param s_filtered: Словарь с ключами-годами и значениями - массивами из зарплат для выбранной профессии
        :param fract: Доли вакансий по городам
        :param cities_s: Средние зарплаты по городам
        """
        self.__salaries_all = s_all
        self.__salaries_filtered = s_filtered
        self.__fraction = fract
        self.__cities_salaries = cities_s
        self.__vacancy = vacancy

        self.__names_ws1 = {
            'A1': 'Год',
            'B1': 'Средняя зарплата',
            'C1': f'Средняя зарплата - {vacancy}',
            'D1': 'Количество вакансий',
            'E1': f'Количество вакансий - {vacancy}',
        }

        self.__names_ws2 = {
            'A1': 'Город',
            'B1': 'Уровень зарплат',
            'D1': 'Город',
            'E1': 'Доля вакансий'
        }
        self.__titles = [
            'Уровень зарплат по годам',
            'Количество вакансий по годам',
            'Уровень зарплат по городам',
            'Доля вакансий по городам',
        ]

    def generate_excel(self):
        """
        Метод для генерации excel отчета
        """
        wb = Workbook()

        ws1 = wb.active
        ws2 = wb.create_sheet('Статистика по городам')

        report.__make_ws1(ws1, self.__salaries_all, self.__salaries_filtered, self.__names_ws1)
        report.__make_ws2(ws2, self.__fraction, self.__cities_salaries, self.__names_ws2)

        wb.save('report.xlsx')

    def generate_image(self):
        """
        Метод для генерации графиков
        """
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        self.__create_bar(
            ax1,
            self.__salaries_all,
            self.__salaries_filtered,
            0,
            ['Средняя з/п', f'З/п {self.__vacancy}'],
            self.__titles[0]
        )

        self.__create_bar(
            ax2,
            self.__salaries_all,
            self.__salaries_filtered,
            1,
            ['Количество вакансий', f'Количество вакансий {self.__vacancy}'],
            self.__titles[1]
        )

        self.__create_barh(ax3, self.__cities_salaries[:10], self.__titles[2])
        self.__create_pie(ax4, self.__fraction[:10], self.__titles[3])

        fig.tight_layout()
        fig.set_size_inches(8, 6)
        fig.set_dpi(300)
        fig.savefig('graph.png', dpi=300)

        plt.show()

    def generate_pdf(self):
        """
        Метод для генерации отчета в формате pdf
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("template.html")
        config = pdfkit.configuration(wkhtmltopdf='/usr/local/bin/wkhtmltopdf')
        path = os.path.abspath('')
        rows_1 = self.__generate_rows_1(self.__salaries_all, self.__salaries_filtered)
        rows_2, rows_3 = self.__generate_rows_23(self.__fraction, self.__cities_salaries)

        context = {
            'path': path,
            'vacancy': self.__vacancy,
            'rows1': rows_1,
            'rows2': rows_2,
            'rows3': rows_3,
        }

        pdf_template = template.render(context)
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})

    @staticmethod
    def __generate_rows_1(s_all: Dict[str, List[int]], s_filtered: Dict[str, List[int]]) -> List[Dict[str, str | int]]:
        """
        Метод для получения списка со статистикой по годам
        :param s_all: Словарь с ключами-годами и значениями - массивами из зарплат
        :param s_filtered: Словарь с ключами-годами и значениями - массивами из зарплат для выбранной профессии
        :return: Массив со статистикой по годам
        """
        rows = []
        for key in s_all.keys():
            row = {
                'year': key,
                'average': s_all[key][0],
                'average_v': s_filtered[key][0],
                'count': s_all[key][1],
                'count_v': s_filtered[key][1],
            }

            rows.append(row)

        return rows

    @staticmethod
    def __generate_rows_23(fract: List[List[float]], cities_s: List[List[int]]) \
            -> Tuple[List[Dict[str, int]], List[Dict[str, float]]]:
        """
        Получение кортежа из списков со статистикой по городам
        :param fract: Доли вакансий по городам
        :param cities_s: Средние зарплаты по городам
        :return: Кортеж из массивов со статистикой по городам
        """
        count = 10
        rows_2 = []
        rows_3 = []

        for i in range(count):
            row = {
                'city': cities_s[i][0],
                'salary': cities_s[i][1]
            }

            rows_2.append(row)

            row = {
                'city': fract[i][0],
                'fraction': str(round(fract[i][1] * 100, 2)) + '%'
            }
            rows_3.append(row)

        return rows_2, rows_3

    @staticmethod
    def __create_bar(
            ax,
            data1: Dict[str, List[int]],
            data2: Dict[str, List[int]],
            index: int,
            legend: List[str],
            title: str
    ):
        """
        Метод для создания столбчатой диаграммы
        :param ax: ax
        :param data1: Словарь с ключами-годами и значениями - массивами из зарплат
        :param data2: Словарь с ключами-годами и значениями - массивами из зарплат
        :param index: Индекс используемых зарплат
        :param legend: Массив для подписей в легенде
        :param title: Название диаграммы
        """
        width = 0.35
        labels_x = data1.keys()
        first = report.__get_data(data1, index)
        second = report.__get_data(data2, index)
        points = range(len(labels_x))

        ax.bar(list(map(lambda x: x - width / 2, points)), first, width, label=legend[0])
        ax.bar(list(map(lambda x: x + width / 2, points)), second, width, label=legend[1])
        ax.set_title(title)
        ax.legend(prop={'size': 8})
        ax.grid(axis='y')

        for label in ax.get_yticklabels():
            label.set_fontsize(8)

        ax.set_xticks(points, labels_x, fontsize=8, rotation=90)

    @staticmethod
    def __create_barh(ax, data: List[List[float]], title: str):
        """
        Метод для создания горизонстальной диаграммы
        :param ax: ax
        :param data: Массив с долями вакансий по городам
        :param title: Название диаграммы
        """
        cities = list(map(lambda x: report.__refactor_label(x[0]), data))
        y_pos = list(range(len(cities)))

        ax.barh(y_pos, list(map(lambda x: x[1], data)), align='center')

        ax.set_yticks(y_pos, labels=cities, fontsize=6)
        ax.invert_yaxis()
        ax.grid(axis='x')

        for label in ax.get_xticklabels():
            label.set_fontsize(8)

        ax.set_title(title)

    @staticmethod
    def __create_pie(ax, data: List[List[float]], title: str):
        """
        Метод для создания круговой диаграммы
        :param ax: ax
        :param data: Массив с долями вакансий
        :param title: Название диаграммы
        """
        cities = list(map(lambda x: x[0], data)) + ['Другие']
        others = 1 - reduce(lambda x, y: x + y[1], data, 0)

        ax.pie(list(map(lambda x: x[1], data)) + [others],
               labels=cities, textprops={'size': 6}, colors=mcolors.BASE_COLORS)

        ax.set_title(title)

    @staticmethod
    def __refactor_label(label: str) -> str:
        """
        Форматирование подписи круговой даиграммы
        :param label: Подпись
        :return: Отформатированная подпись
        """
        spaces = re.compile('\s+')
        line = re.compile('-+')

        label = re.sub(spaces, '\n', label)
        return re.sub(line, '-\n', label)

    @staticmethod
    def __get_data(data: Dict[str, List[int]], i: int) -> List[int]:
        """
        Метод для получения массива с данными для графика
        :param data: Словарь с ключами-годами и значениями - массивами из зарплат
        :param i: Индекс нужного элемента в массиве
        :return: Массив с данными для графика
        """
        return list(map(lambda x: x[i], data.values()))

    @staticmethod
    def __make_ws1(
            ws,
            s_all: Dict[str, List[int]],
            s_filtered: Dict[str, List[int]],
            title: Dict[str, str]
    ):
        """
        Метод для заполнения первого листа excel отчета
        :param ws: Лист
        :param s_all: Словарь с ключами-годами и значениями - массивами из зарплат
        :param s_filtered: Словарь с ключами-годами и значениями - массивами из зарплат для данной профессии
        :param title: Название листа
        """
        ws.title = 'Статистика по годам'
        report.__create_title(ws, title)

        for key in s_all.keys():
            ws.append([
                int(key),
                s_all[key][0],
                s_filtered[key][0],
                s_all[key][1],
                s_filtered[key][1],
            ])

        report.__set_border(ws, f'A1:E{len(s_all) + 1}')
        report.__refactor_rows(ws)

    @staticmethod
    def __make_ws2(
            ws,
            fract: List[List[float]],
            cities_s: List[List[int]],
            title: Dict[str, str]
    ):
        """
        Метод для заполнения второго листа excel отчета
        :param ws: Лист
        :param fract: Массив с долями вакансий по городам
        :param cities_s: Массив с уровнем зарплат по городам
        :param title: Название листа
        """
        report.__create_title(ws, title)
        count = 10

        for i in range(count):
            row = []

            row += [cities_s[i][0], cities_s[i][1]] if len(cities_s) >= i + 1 else ['', '']
            row += ['']
            row += [fract[i][0], fract[i][1]] if len(fract) >= i + 1 else ['', '']

            ws.append(row)

        report.__add_percentage(ws, count, 'E')
        report.__set_border(ws, f'A1:B{count + 1}')
        report.__set_border(ws, f'D1:E{count + 1}')
        report.__refactor_rows(ws)

    @staticmethod
    def __add_percentage(ws, count: int, column: str):
        """
        Добавить процентный формат данных определенному столбцу
        :param ws: Лист
        :param count: Количество строк
        :param column: Колонка
        """
        for i in range(2, count + 2):
            ws[f'{column}{i}'].number_format = FORMAT_PERCENTAGE_00

    @staticmethod
    def __set_border(ws, cell_range):
        """
        Добавить границы диапазону ячеек
        :param ws: Лист
        :param cell_range: Диапазон ячеек
        """
        line = Side(border_style="thin", color="000000")
        border = Border(top=line, left=line, right=line, bottom=line)

        for row in ws[cell_range]:
            for cell in row:
                cell.border = border

    @staticmethod
    def __create_title(ws, title: Dict[str, str]):
        """
        Добавить названия для стобцов
        :param ws: Лист
        :param title: Названия столбцов по ячейкам
        """
        font = Font(bold=True)

        for key, value in title.items():
            ws[key] = value
            ws[key].font = font

    @staticmethod
    def __refactor_rows(ws):
        """
        Установить минимально возможную ширину для ячеек на листе
        :param ws: Лист
        """
        for i, col in enumerate(ws.iter_cols()):
            l = 0
            for cell in col:
                v = cell.value if cell.value is not None else ''
                l = max(l, len(str(v)))

            ws.column_dimensions[get_column_letter(i + 1)].width = l + 3 if l != 0 else 0


connect = InputConnect()
connect.read_console()

dataset = DataSet(connect.file_name)

salaries_all = dataset.get_vacancies_years()
salaries_filtered = dataset.get_vacancies_years(lambda x: x.is_suitible(connect.vacancy))
fraction, cities_salaries = dataset.get_vacancies_cities()

rep = report(connect.vacancy,
             salaries_all,
             salaries_filtered,
             fraction,
             cities_salaries
             )

if connect.method.lower() == 'статистика':
    connect.write_console(salaries_all, salaries_filtered, fraction, cities_salaries)
    rep.generate_excel()
else:
    rep.generate_excel()
    rep.generate_image()
    rep.generate_pdf()
